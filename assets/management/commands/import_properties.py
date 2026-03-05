"""
Import properties from Stan Gribble property ownership Excel spreadsheet.

Usage:
    python manage.py import_properties /path/to/spreadsheet.xlsx
    python manage.py import_properties /path/to/spreadsheet.xlsx --dry-run
    python manage.py import_properties /path/to/spreadsheet.xlsx --update
"""
import datetime
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand

from assets.models import PropertyOwnership, RealEstate
from stakeholders.models import Stakeholder

OWNER_NAME = "314SG, LLC"
MANAGER_NAME = "CAP 2, LLC / Equitas Investments, LLC"

FINANCIAL_FIELDS = [
    "sold_date", "unreturned_capital", "total_unreturned_capital",
    "loan_balance_snapshot", "equity", "deferred_gain",
    "monthly_income", "monthly_accrued_income", "total_accrued_pref_return",
    "income_source",
]

FINANCIAL_LABELS = {
    "unreturned_capital": "Unreturned Capital",
    "total_unreturned_capital": "Total Unreturned Capital",
    "loan_balance_snapshot": "Loan Balance",
    "equity": "Equity",
    "deferred_gain": "Deferred Gain",
    "monthly_income": "Monthly Income",
    "monthly_accrued_income": "Accrued Income",
    "total_accrued_pref_return": "Accrued Pref Return",
    "income_source": "Income Source",
}


class Command(BaseCommand):
    help = "Import properties from Excel spreadsheet into RealEstate model"

    def add_arguments(self, parser):
        parser.add_argument("file", help="Path to the .xlsx file")
        parser.add_argument("--dry-run", action="store_true",
                            help="Preview what would be imported without saving")
        parser.add_argument("--update", action="store_true",
                            help="Update existing properties with financial data instead of skipping")
        parser.add_argument("--sheet", type=str, default=None,
                            help="Sheet name to read (default: first sheet)")

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write("openpyxl is required: pip install openpyxl")
            return

        filepath = options["file"]
        dry_run = options["dry_run"]
        update = options["update"]

        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet_name = options["sheet"] or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            self.stderr.write(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            return
        ws = wb[sheet_name]
        self.stdout.write(f"Reading sheet: {sheet_name}")

        # Parse the hierarchical spreadsheet structure
        properties = self._parse_spreadsheet(ws)

        mode = "DRY RUN" if dry_run else ("IMPORT + UPDATE" if update else "IMPORT")
        self.stdout.write(self.style.WARNING(
            f"\n=== {mode} — {len(properties)} properties found ===\n"
        ))

        # Ensure all stakeholders exist before processing properties
        self.stdout.write(self.style.WARNING("--- Stakeholders ---"))
        stakeholder_map, stakeholders_created = self._ensure_stakeholders(
            properties, dry_run=dry_run
        )
        self.stdout.write("")

        created_count = 0
        updated_count = 0
        skipped_count = 0
        ownership_count = 0

        for prop in properties:
            existing = RealEstate.objects.filter(address=prop["address"]).first()

            if existing and not update:
                self.stdout.write(self.style.WARNING(
                    f"  SKIP (exists): {prop['name']}"
                ))
                ownership_count += self._create_ownership_records(
                    existing, prop, stakeholder_map, dry_run=dry_run
                )
                skipped_count += 1
                continue

            if existing and update:
                # Update financial fields on existing property
                changed_fields = []
                for field in FINANCIAL_FIELDS:
                    new_val = prop.get(field)
                    if new_val is not None and new_val != "":
                        old_val = getattr(existing, field)
                        if old_val != new_val:
                            setattr(existing, field, new_val)
                            changed_fields.append(field)
                # Also update status for sold properties
                if prop["status"] == "sold" and existing.status != "sold":
                    existing.status = "sold"
                    changed_fields.append("status")

                if not changed_fields:
                    self.stdout.write(f"  NO CHANGE: {existing.name}")
                    ownership_count += self._create_ownership_records(
                        existing, prop, stakeholder_map, dry_run=dry_run
                    )
                    skipped_count += 1
                    continue

                if dry_run:
                    self.stdout.write(
                        f"  WOULD UPDATE: {existing.name} (pk={existing.pk})\n"
                        f"    Fields: {', '.join(changed_fields)}"
                    )
                else:
                    existing.save()
                    self.stdout.write(self.style.SUCCESS(
                        f"  UPDATED: {existing.name} (pk={existing.pk}) — {', '.join(changed_fields)}"
                    ))
                ownership_count += self._create_ownership_records(
                    existing, prop, stakeholder_map, dry_run=dry_run
                )
                updated_count += 1
                continue

            # New property — create it
            if dry_run:
                val_str = f"${prop['estimated_value']:,.2f}" if prop["estimated_value"] else "N/A"
                self.stdout.write(
                    f"  WOULD CREATE: {prop['name']}\n"
                    f"    Address: {prop['address']}\n"
                    f"    Status: {prop['status']}  |  Value: {val_str}"
                )
                if prop.get("sold_date"):
                    self.stdout.write(f"    Sold: {prop['sold_date']}")
                for field_key, field_label in FINANCIAL_LABELS.items():
                    val = prop.get(field_key)
                    if val:
                        if isinstance(val, Decimal):
                            self.stdout.write(f"    {field_label}: ${val:,.2f}")
                        else:
                            self.stdout.write(f"    {field_label}: {val}")
                self.stdout.write(f"    Ownership links:")
                ownership_count += self._create_ownership_records(
                    None, prop, stakeholder_map, dry_run=True
                )
                self.stdout.write("")
            else:
                re = RealEstate.objects.create(
                    name=prop["name"],
                    tenant=prop.get("tenant", ""),
                    address=prop["address"],
                    property_type=prop.get("property_type", ""),
                    estimated_value=prop.get("estimated_value"),
                    acquisition_date=prop.get("acquisition_date"),
                    status=prop["status"],
                    notes_text=prop.get("notes_text", ""),
                    jurisdiction=prop.get("jurisdiction", ""),
                    sold_date=prop.get("sold_date"),
                    unreturned_capital=prop.get("unreturned_capital"),
                    total_unreturned_capital=prop.get("total_unreturned_capital"),
                    loan_balance_snapshot=prop.get("loan_balance_snapshot"),
                    equity=prop.get("equity"),
                    deferred_gain=prop.get("deferred_gain"),
                    monthly_income=prop.get("monthly_income"),
                    monthly_accrued_income=prop.get("monthly_accrued_income"),
                    total_accrued_pref_return=prop.get("total_accrued_pref_return"),
                    income_source=prop.get("income_source", ""),
                )
                self.stdout.write(self.style.SUCCESS(f"  CREATED: {re.name} (pk={re.pk})"))
                ownership_count += self._create_ownership_records(
                    re, prop, stakeholder_map, dry_run=False
                )
            created_count += 1

        self.stdout.write("")
        parts = []
        if created_count:
            parts.append(f"{created_count} properties {'would be created' if dry_run else 'created'}")
        if updated_count:
            parts.append(f"{updated_count} properties {'would be updated' if dry_run else 'updated'}")
        if skipped_count:
            parts.append(f"{skipped_count} properties skipped")
        if stakeholders_created:
            parts.append(f"{stakeholders_created} stakeholders {'would be created' if dry_run else 'created'}")
        if ownership_count:
            parts.append(f"{ownership_count} ownership links {'would be created' if dry_run else 'created'}")
        self.stdout.write(self.style.SUCCESS(f"Done: {', '.join(parts)}"))

    def _parse_spreadsheet(self, ws):
        """Parse the hierarchical spreadsheet into a flat list of property dicts."""
        properties = []
        current_entity = None
        current_own_pct = None
        current_entity_num = None

        # Read all rows (skip header rows 1-6)
        rows = list(ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True))

        for i, row in enumerate(rows):
            col_a = row[0]  # Entity number
            col_c = row[2]  # Entity name
            col_d = row[3]  # Tenant
            col_f = row[5]  # Own %
            col_g = row[6]  # Address
            col_i = row[8]  # Purchased
            col_k = row[10]  # Status
            col_l = row[11]  # Sold date
            col_n = row[13]  # Unreturned capital (314SG)
            col_o = row[14]  # Total unreturned capital
            col_p = row[15]  # Fair Market Value
            col_q = row[16]  # Loan balance
            col_r = row[17]  # Equity
            col_s = row[18]  # Deferred gain
            col_t = row[19]  # Monthly income
            col_u = row[20]  # Monthly accrued income
            col_v = row[21]  # Total accrued pref return
            col_w = row[22]  # Income source
            col_x = row[23]  # Notes

            # Detect entity group header rows (have a number in col A)
            if col_a is not None and isinstance(col_a, (int, float)):
                current_entity_num = int(col_a)
                current_entity = col_c if col_c else None
                # Own % from group header
                if col_f is not None and isinstance(col_f, (int, float)):
                    current_own_pct = col_f
                else:
                    current_own_pct = None
                continue

            # Detect entity name rows (col C has entity name, no tenant in col D)
            if col_c and not col_d and not col_g:
                current_entity = col_c
                continue

            # Detect property rows (must have tenant in col D and address in col G)
            if not col_d or not col_g:
                continue

            # Get status
            status_raw = str(col_k).strip() if col_k else ""

            # Map spreadsheet status to model status
            status_upper = status_raw.upper()
            if status_upper == "ACTIVE":
                model_status = "owned"
            elif status_upper == "SOLD":
                model_status = "sold"
            elif status_upper == "TIC'D":
                model_status = "sold"
            elif status_upper == "DEED BACK TO SGAS":
                model_status = "sold"
            else:
                # Skip unknown statuses
                continue

            # Build property name: "Tenant - Street Address"
            address_full = str(col_g).strip()
            street = address_full.split(",")[0].strip()
            tenant = str(col_d).strip()
            name = f"{tenant} - {street}"

            # Parse acquisition date
            acq_date = None
            if isinstance(col_i, datetime.datetime):
                acq_date = col_i.date()
            elif isinstance(col_i, datetime.date):
                acq_date = col_i

            # Parse sold date
            sold_date = None
            if isinstance(col_l, datetime.datetime):
                sold_date = col_l.date()
            elif isinstance(col_l, datetime.date):
                sold_date = col_l

            # Parse fair market value
            fmv = self._parse_decimal(col_p)

            # Parse financial fields
            unreturned_capital = self._parse_decimal(col_n)
            total_unreturned_capital = self._parse_decimal(col_o)
            loan_balance_snapshot = self._parse_decimal(col_q)
            equity = self._parse_decimal(col_r)
            deferred_gain = self._parse_decimal(col_s)
            monthly_income = self._parse_decimal(col_t)
            monthly_accrued_income = self._parse_decimal(col_u)
            total_accrued_pref_return = self._parse_decimal(col_v)

            # Income source (string)
            income_source = str(col_w).strip() if col_w else ""

            # Determine ownership percentage for this property
            own_pct = current_own_pct
            if col_f is not None and isinstance(col_f, (int, float)):
                own_pct = col_f

            # Build notes from entity info + spreadsheet notes
            notes_parts = []
            if current_entity:
                notes_parts.append(f"Entity: {current_entity}")
            if own_pct is not None:
                pct_display = f"{own_pct * 100:.2f}%" if own_pct <= 1 else f"{own_pct:.2f}%"
                notes_parts.append(f"314SG Ownership: {pct_display}")

            # Collect notes from this row and continuation rows
            if col_x:
                note_text = str(col_x).strip()
                # Look ahead for continuation note rows
                for j in range(i + 1, min(i + 5, len(rows))):
                    next_row = rows[j]
                    # Continuation rows have notes in col X but no tenant/address
                    if next_row[3] is None and next_row[6] is None and next_row[23]:
                        note_text += " " + str(next_row[23]).strip()
                    else:
                        break
                notes_parts.append(note_text)

            # Extract jurisdiction (state) from address
            jurisdiction = self._extract_state(address_full)

            # Compute ownership percentage as a Decimal (0-100 scale)
            if own_pct is not None:
                if own_pct <= 1:
                    ownership_pct = Decimal(str(own_pct * 100)).quantize(Decimal("0.01"))
                else:
                    ownership_pct = Decimal(str(own_pct)).quantize(Decimal("0.01"))
            else:
                ownership_pct = None

            properties.append({
                "name": name,
                "tenant": tenant,
                "address": address_full,
                "property_type": "Commercial",
                "estimated_value": fmv,
                "acquisition_date": acq_date,
                "status": model_status,
                "jurisdiction": jurisdiction,
                "notes_text": "\n".join(notes_parts),
                "sold_date": sold_date,
                "unreturned_capital": unreturned_capital,
                "total_unreturned_capital": total_unreturned_capital,
                "loan_balance_snapshot": loan_balance_snapshot,
                "equity": equity,
                "deferred_gain": deferred_gain,
                "monthly_income": monthly_income,
                "monthly_accrued_income": monthly_accrued_income,
                "total_accrued_pref_return": total_accrued_pref_return,
                "income_source": income_source,
                "entity_llc": current_entity,
                "ownership_pct": ownership_pct,
            })

        return properties

    def _ensure_stakeholders(self, properties, dry_run=False):
        """Create or look up all stakeholders needed for ownership records.

        Returns a dict mapping name → Stakeholder instance (empty in dry-run).
        """
        # Collect unique entity LLC names from parsed properties
        entity_names = {p["entity_llc"] for p in properties if p.get("entity_llc")}
        all_names = [OWNER_NAME, MANAGER_NAME] + sorted(entity_names)

        stakeholder_map = {}
        created_count = 0

        for name in all_names:
            if dry_run:
                exists = Stakeholder.objects.filter(name=name).exists()
                if exists:
                    self.stdout.write(f"  Stakeholder exists: {name}")
                else:
                    self.stdout.write(self.style.NOTICE(
                        f"  WOULD CREATE stakeholder: {name}"
                    ))
                    created_count += 1
            else:
                stakeholder, created = Stakeholder.objects.get_or_create(
                    name=name,
                    defaults={"entity_type": "business_partner"},
                )
                stakeholder_map[name] = stakeholder
                if created:
                    self.stdout.write(self.style.SUCCESS(
                        f"  Created stakeholder: {name} (pk={stakeholder.pk})"
                    ))
                    created_count += 1
                else:
                    self.stdout.write(f"  Stakeholder exists: {name} (pk={stakeholder.pk})")

        # Set parent_organization on entity LLCs → 314SG, LLC
        owner = stakeholder_map.get(OWNER_NAME)
        if owner and not dry_run:
            for name, s in stakeholder_map.items():
                if name != OWNER_NAME and not s.parent_organization:
                    s.parent_organization = owner
                    s.save(update_fields=["parent_organization"])

        return stakeholder_map, created_count

    def _create_ownership_records(self, real_estate, prop, stakeholder_map, dry_run=False):
        """Create PropertyOwnership records linking a property to its stakeholders.

        Returns the number of ownership records created (or would-be-created).
        """
        records_to_create = []

        # 1. 314SG, LLC as Owner with ownership percentage
        records_to_create.append({
            "stakeholder_name": OWNER_NAME,
            "role": "Owner",
            "percentage": prop.get("ownership_pct"),
        })

        # 2. Manager
        records_to_create.append({
            "stakeholder_name": MANAGER_NAME,
            "role": "Manager",
            "percentage": None,
        })

        # 3. Entity LLC as Holding Entity (if present)
        if prop.get("entity_llc"):
            records_to_create.append({
                "stakeholder_name": prop["entity_llc"],
                "role": "Holding Entity",
                "percentage": None,
            })

        created_count = 0
        for rec in records_to_create:
            if dry_run:
                pct_str = f" ({rec['percentage']}%)" if rec["percentage"] else ""
                self.stdout.write(
                    f"      -> {rec['stakeholder_name']} as {rec['role']}{pct_str}"
                )
                created_count += 1
            else:
                stakeholder = stakeholder_map[rec["stakeholder_name"]]
                _, created = PropertyOwnership.objects.get_or_create(
                    property=real_estate,
                    stakeholder=stakeholder,
                    defaults={
                        "ownership_percentage": rec["percentage"],
                        "role": rec["role"],
                    },
                )
                if created:
                    created_count += 1

        return created_count

    def _parse_decimal(self, value):
        """Parse a cell value into a Decimal, returning None on failure."""
        if value is None:
            return None
        if isinstance(value, (int, float)):
            try:
                return Decimal(str(value)).quantize(Decimal("0.01"))
            except (InvalidOperation, ValueError):
                return None
        return None

    def _extract_state(self, address):
        """Try to extract state abbreviation from address string."""
        import re
        # Match 2-letter state code before zip
        match = re.search(r',\s*([A-Z]{2})\s+\d{5}', address)
        if match:
            return match.group(1)
        # Try without zip
        match = re.search(r',\s*([A-Z]{2})\s*$', address)
        if match:
            return match.group(1)
        return ""
